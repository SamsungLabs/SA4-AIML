#!/usr/bin/env python

import tensorflow as tf
import numpy as np
import argparse
import os
import cv2
import openpyxl
from openpyxl.styles import Font
from termcolor import colored
from pathlib import Path
from datetime import datetime

parser = argparse.ArgumentParser()
parser._action_groups.pop()
required = parser.add_argument_group('required arguments')
optional = parser.add_argument_group('optional arguments')
required.add_argument('--model_path', type = Path, required = True, help = 'path to the Frozen model (.pb)')
required.add_argument('--image_dir', type = Path, required = True, help = 'path to the test images directory')
optional.add_argument('--split_layer', type = int, default = 12, choices = range(1, 28), 
                      help = 'layer at which the model should be splitted. Accepted range of values = [1, 27]')
optional.add_argument('--output_dir', type = Path, default = "./output", help = 'path to the output directory')

input_name = 'image'
input_dim = [257, 257, 3]
tensor_name_prefix = "MobilenetV1/Relu6"
tensor_out = ['heatmap_2:0', 'offset_2:0', 'displacement_fwd_2:0', 'displacement_bwd_2:0']

model_dict = {}
model_dict[0] = [129, 129, 32]
model_dict[1] = [129, 129, 32]
model_dict[2] = [129, 129, 64]
model_dict[3] = [65, 65, 64]
model_dict[4] = [65, 65, 128]
model_dict[5] = [65, 65, 128]
model_dict[6] = [65, 65, 128]
model_dict[7] = [33, 33, 128]
model_dict[8] = [33, 33, 256]
model_dict[9] = [33, 33, 256]
model_dict[10] = [33, 33, 256]
model_dict[11] = [17, 17, 256]
model_dict[12] = [17, 17, 512]
model_dict[13] = [17, 17, 512]
model_dict[14] = [17, 17, 512]
model_dict[15] = [17, 17, 512]
model_dict[16] = [17, 17, 512]
model_dict[17] = [17, 17, 512]
model_dict[18] = [17, 17, 512]
model_dict[19] = [17, 17, 512]
model_dict[20] = [17, 17, 512]
model_dict[21] = [17, 17, 512]
model_dict[22] = [17, 17, 512]
model_dict[23] = [17, 17, 512]
model_dict[24] = [17, 17, 1024]
model_dict[25] = [17, 17, 1024]
model_dict[26] = [17, 17, 1024]

part_names = [
    "nose", "leftEye", "rightEye", "leftEar", "rightEar", "leftShoulder",
    "rightShoulder", "leftElbow", "rightElbow", "leftWrist", "rightWrist",
    "leftHip", "rightHip", "leftKnee", "rightKnee", "leftAnkle", "rightAnkle"
]

output_stride = 32

def get_graph(filename):
    graph_def = tf.compat.v1.GraphDef() 
    with tf.io.gfile.GFile(filename, 'rb') as f:
        graph_def.ParseFromString(f.read())
        
    with tf.Graph().as_default() as graph:
        new_input = tf.compat.v1.placeholder(
            np.float32, shape = [None, input_dim[0], input_dim[1], input_dim[2]], name = 'new_input')
        tf.import_graph_def(
            graph_def, {input_name: new_input}, name = '')
    return graph_def, graph

def generate_tflite(filename, output_dir):
    graph = get_graph(filename)[1]
    output_file = os.path.join(output_dir, "full_model.tflite")      
    with tf.compat.v1.Session(graph = graph) as sess:
        _in = sess.graph.get_tensor_by_name('new_input:0')
        _out = []
        for tensor in tensor_out:
            _out.append(sess.graph.get_tensor_by_name(tensor))
        
        full_model = tf.compat.v1.lite.TFLiteConverter.from_session(
            sess, [_in], _out).convert()
        open(output_file, "wb").write(full_model)
    return output_file

def split_model(filename, layer, output_dir):
    partitioned_node = tensor_name_prefix
    if layer > 0:
        partitioned_node = partitioned_node + '_' + str(layer)
        
    graph_def, graph_1 = get_graph(filename)
    
    with tf.Graph().as_default() as graph_2:
        input_middle = tf.compat.v1.placeholder(
            np.float32, shape = [None, model_dict[layer][0], model_dict[layer][1], model_dict[layer][2]], name = 'input_middle')
        tf.import_graph_def(
            graph_def, {partitioned_node: input_middle}, name = '')
    
    # Make Models
    partitioned_tensor = partitioned_node + ':0'
    output_file_1 = os.path.join(output_dir, "first_model.tflite")
    with tf.compat.v1.Session(graph = graph_1) as sess_1:
        device_in = sess_1.graph.get_tensor_by_name('new_input:0')
        device_out = sess_1.graph.get_tensor_by_name(partitioned_tensor)
        
        device_model = tf.compat.v1.lite.TFLiteConverter.from_session(
            sess_1, [device_in], [device_out]).convert()
        open(output_file_1, "wb").write(device_model)
    
    output_file_2 = os.path.join(output_dir, "second_model.tflite")
    with tf.compat.v1.Session(graph = graph_2) as sess_2:
        edge_in = sess_2.graph.get_tensor_by_name('input_middle:0')
        edge_out = []
        for tensor in tensor_out:
            edge_out.append(sess_2.graph.get_tensor_by_name(tensor))
        
        edge_model = tf.compat.v1.lite.TFLiteConverter.from_session(
            sess_2, [edge_in], edge_out).convert()
        open(output_file_2, "wb").write(edge_model)
    return output_file_1, output_file_2

def get_heatmap_scores(heatmaps_result):
    height, width, depth = heatmaps_result.shape
    reshaped_heatmap = np.reshape(heatmaps_result, [height * width, depth])
    coords = np.argmax(reshaped_heatmap, axis=0)
    y_coords = coords // width
    x_coords = coords % width
    return np.concatenate([np.expand_dims(y_coords, 1), np.expand_dims(x_coords, 1)], axis=1)

def get_points_confidence(heatmaps, coords):
    result = []
    for keypoint in range(len(part_names)):
        # Get max value of heatmap for each keypoint
        result.append(heatmaps[coords[keypoint, 0],coords[keypoint, 1], keypoint])
    return result

def get_offset_vectors(coords, offsets_result):
    result = []
    for keypoint in range(len(part_names)):
        heatmap_y = coords[keypoint, 0]
        heatmap_x = coords[keypoint, 1]

        offset_y = offsets_result[heatmap_y, heatmap_x, keypoint]
        offset_x = offsets_result[heatmap_y, heatmap_x, keypoint + len(part_names)]

        result.append([offset_y, offset_x])
    # print(result)
    return result

def get_offset_points(coords, offsets_result, output_stride=output_stride):
    offset_vectors = get_offset_vectors(coords, offsets_result)
    scaled_heatmap = coords * output_stride
    return scaled_heatmap + offset_vectors

def decode_single_pose(heatmaps, offsets, img_width, img_height):
    width_factor = img_width / input_dim[0]
    height_factor = img_height / input_dim[1]
    poses = []
    heatmaps_coords = get_heatmap_scores(heatmaps)
    offset_points = get_offset_points(heatmaps_coords, offsets, output_stride)
    keypoint_confidence = get_points_confidence(heatmaps, heatmaps_coords)

    keypoints = [{
        "position": {
            "y": offset_points[keypoint, 0] * height_factor,
            "x": offset_points[keypoint, 1] * width_factor
        },
        "part": part_names[keypoint],
        "score": score
    } for keypoint, score in enumerate(keypoint_confidence)]

    poses.append({"keypoints": keypoints, \
                  "score": (sum(keypoint_confidence) / len(keypoint_confidence))})
    #print(poses)
    return poses

def generate_input_data(image, tensor_input_details):
    img = cv2.resize(image, (tensor_input_details[0]['shape'][1], tensor_input_details[0]['shape'][2]), interpolation = cv2.INTER_CUBIC).astype(np.float32)
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB).astype(np.float32)
    img = img * (2.0 / 255.0) - 1.0
    input_data = np.expand_dims(img, axis = 0)
    return input_data

def get_output(interpreter, tensor_output_details):
    heatmaps = np.array(interpreter.get_tensor(tensor_output_details[0]['index']))
    offsets = np.array(interpreter.get_tensor(tensor_output_details[1]['index']))
    displacement_fwd = np.array(interpreter.get_tensor(tensor_output_details[2]['index']))
    displacement_bwd = np.array(interpreter.get_tensor(tensor_output_details[3]['index']))
    heatmaps = np.squeeze(heatmaps)
    offsets = np.squeeze(offsets)
    displacement_fwd = np.squeeze(displacement_fwd)
    displacement_bwd = np.squeeze(displacement_bwd)
    return heatmaps, offsets, displacement_fwd, displacement_bwd

def run_full_model_test(files, model_path):
    interpreter = tf.lite.Interpreter(model_path = model_path)
    interpreter.allocate_tensors()
    input_details = interpreter.get_input_details()
    output_details = interpreter.get_output_details()
    
    result = []
    start_time = datetime.now()
    for f in files:
        original_img = cv2.imread(f)
        height, width, channels = original_img.shape
        input_data = generate_input_data(original_img, input_details)
        
        interpreter.set_tensor(input_details[0]['index'], input_data)
        interpreter.invoke()
        
        heatmaps, offsets = get_output(interpreter, output_details)[0:2]
        poses = decode_single_pose(heatmaps, offsets, width, height)
        result.append(poses)
    end_time = datetime.now()
    time_difference = end_time - start_time
    average_time = (1000.0 * time_difference.total_seconds()) / len(files)
    return result, average_time

def run_split_model_test(files, first_model_path, second_model_path):
    interpreter_1 = tf.lite.Interpreter(model_path = first_model_path)
    interpreter_1.allocate_tensors()
    input_details_1 = interpreter_1.get_input_details()
    output_details_1 = interpreter_1.get_output_details()
    
    interpreter_2 = tf.lite.Interpreter(model_path = second_model_path)
    interpreter_2.allocate_tensors()
    input_details_2 = interpreter_2.get_input_details()
    output_details_2 = interpreter_2.get_output_details()
    
    result = []
    start_time = datetime.now()
    for f in files:
        original_img = cv2.imread(f)
        height, width, channels = original_img.shape
        input_data = generate_input_data(original_img, input_details_1)
        
        interpreter_1.set_tensor(input_details_1[0]['index'], input_data)
        interpreter_1.invoke()
        
        partial_output = np.array(interpreter_1.get_tensor(output_details_1[0]['index']))
        interpreter_2.set_tensor(input_details_2[0]['index'], partial_output)
        interpreter_2.invoke()
        
        heatmaps, offsets = get_output(interpreter_2, output_details_2)[0:2]
        poses = decode_single_pose(heatmaps, offsets, width, height)
        result.append(poses)
    end_time = datetime.now()
    time_difference = end_time - start_time
    average_time = (1000.0 * time_difference.total_seconds()) / len(files)
    return result, average_time

def write_to_excel(tests_list, file_names, output_file, index, title):
    if os.path.isfile(output_file):
        wb = openpyxl.load_workbook(output_file)
    else:
        wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"]) 
    if title not in wb.sheetnames:
        wb.create_sheet(index = index, title = title)
    sheet = wb[title]
    sheet.merge_cells(start_row = 1, start_column = 1, end_row = 2, end_column = 1)
    cell = sheet.cell(row = 1, column = 1)
    cell.value = "Images"
    cell.font = Font(bold = True)
    
    col = 2
    for part in part_names:
        sheet.merge_cells(start_row = 1, start_column = col, end_row = 1, end_column = col + 2)
        cell = sheet.cell(row = 1, column = col)
        cell.value = "Part: " + part
        cell.font = Font(bold = True)
        cell = sheet.cell(row = 2, column = col)
        cell.value = "Position: x"
        cell.font = Font(bold = True)
        cell = sheet.cell(row = 2, column = col + 1)
        cell.value = "Position: y"
        cell.font = Font(bold = True)
        cell = sheet.cell(row = 2, column = col + 2)
        cell.value = "Confidence score"
        cell.font = Font(bold = True)
        col += 3
    
    num_tests = len(tests_list)
    row = 3
    for test in range(num_tests):
        file = file_names[test]
        cell = sheet.cell(row = row, column = 1)
        cell.value = file
        pose = tests_list[test][0]["keypoints"]
        col = 2
        for keypoint in pose:
            x = keypoint["position"]["x"]
            cell = sheet.cell(row = row, column = col)
            cell.value = x
            y = keypoint["position"]["y"]
            cell = sheet.cell(row = row, column = col + 1)
            cell.value = y
            score = keypoint["score"]
            cell = sheet.cell(row = row, column = col + 2)
            cell.value = score
            col += 3
        row += 1
        
    wb.save(output_file)
        
        
def compare(val_1, val_2):
    diff = abs(val_1 - val_2)
    if diff > 0.001:
        return False
    return True
    
if __name__ == "__main__":
    args = parser.parse_args()
    model_path = args.model_path
    if not os.path.exists(model_path):
        raise Exception("Invalid model file")
    
    image_dir = args.image_dir
    if not os.path.exists(image_dir):
        raise Exception("Invalid image directory")
    
    file_names = [f.path for f in os.scandir(image_dir) if f.is_file() and f.path.endswith(('.png', '.jpg'))]
    
    output_dir = args.output_dir
    full_model_path = generate_tflite(model_path, output_dir)
    print(colored("Completed TFLite full model generation", "green"))
    print("\n")
    
    split_layer = args.split_layer
    split_layer -= 1
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    first_model_path, second_model_path = split_model(model_path, split_layer, output_dir)
    print(colored("Completed model splitting", "green"))
    print("\n")
    
    output_file = os.path.join(output_dir, "result.xlsx")
    print("Starting full model inference test")
    result_1, avg_time = run_full_model_test(file_names, full_model_path)
    write_to_excel(result_1, file_names, output_file, 0, "Full model")
    print(colored("Completed full model inference test. Avg. time taken per input: " + str(avg_time) + "ms", "green"))
    print("\n")
    
    print("Starting split model inference test")
    result_2, avg_time = run_split_model_test(file_names, first_model_path, second_model_path)
    write_to_excel(result_2, file_names, output_file, 1, "Split models")
    print(colored("Completed split model inference test. Avg. time taken per input: " + str(avg_time) + "ms", "green"))
    print("\n")
    
    print(colored("Wrote results to file: " + output_file, "green"))
    print("\n")
    
    num_keypoints = len(part_names)
    num_tests = len(file_names)
    for test in range(num_tests):
        pose_1 = result_1[test][0]["keypoints"]
        pose_2 = result_2[test][0]["keypoints"]
        is_pass = True
        for i in range(num_keypoints):
            x_coord_1 = pose_1[i]["position"]["x"]
            x_coord_2 = pose_2[i]["position"]["x"]
            is_pass = is_pass & compare(x_coord_1, x_coord_2)
                
            y_coord_1 = pose_1[i]["position"]["y"]
            y_coord_2 = pose_2[i]["position"]["y"]
            is_pass = is_pass & compare(y_coord_1, y_coord_2)
            
            score_1 = pose_1[i]["score"]
            score_2 = pose_2[i]["score"]
            is_pass = is_pass & compare(score_1, score_2)
        
        if is_pass == False:
            print(colored("FAILED:", "red") + " Inference results don't match for input: " + file_names[test])
        else:
            print(colored("PASSED:", "green") + " Inference results match for input: " + file_names[test])
